import openpyxl
wk=openpyxl.load_workbook('数据.xlsx')
sheet=wk.active
row=sheet.max_row
cols=sheet.max_column


list=[]
for i in range(1,row+1):
    color=sheet.cell(i,2).value
    list.append(color)
# for item in list:
#     print(item)

dic_color={}
for item in list:
    dic_color[item]=0


#for item in dic_color:
    #print(item,dic_color[item])


for item in list:
    for color in dic_color:
        if item==color:
            dic_color[color]+=1
            break

for item in dic_color:
    print(item,dic_color[item])